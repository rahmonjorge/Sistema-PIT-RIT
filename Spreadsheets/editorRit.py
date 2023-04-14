from openpyxl.drawing.image import Image 
from openpyxl.styles import NamedStyle
import grpc
from database import sheets_pb2, sheets_pb2_grpc 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

checked = Image("checked.png") 
unchecked = Image("unchecked.png")

wb = load_workbook('teste checkbox.xlsx')
ws = wb.active
ws.title = 'Rit'

channel = grpc.insecure_channel('26.52.183.15:6924')
stub = sheets_pb2_grpc.SpreadsheetServiceStub(channel)

getUser1 = sheets_pb2.GetUserRequest(id='641f3e1adce65c4ec0b00266')
    
get = stub.GetUser(getUser1)

getSheet1 = sheets_pb2.GetSheetRequest(userId='641f3e1adce65c4ec0b00266',type='rit', ano=2023)

sheet = stub.GetSheet(getSheet1)


def dadosCabecalho():
    ws['C3'].value = (get.name)
    ws['C4'].value = (get.siape)
    ws['C5'].value = (get.departamento)
    ws['C6'].value = (get.email)
    ws['C7'].value = (get.vinculo)
    ws['C8'].value = (get.regime)
    ws['C9'].value = (get.reducao)

def ritAulas():
    ws['J13'].value = (get.ch_grad)
    ws['J14'].value = (get.ch_pos)

def subTotais():
    ws['J35'].value = (get.ch_ensino)
    ws['J68'].value = (get.ch_pesquisa)
    ws['J90'].value = (get.ch_extensao)
    ws['J105'].value = (get.ch_adm)


def ritAtividadesEnsino():
    for i in range(0,15):
        checked = Image("checked.png") 
        checked.height = 30
        checked.width = 30 
        unchecked = Image("unchecked.png")
        unchecked.height = 30
        unchecked.width = 30
        celula = 'J' + str(i+19)
        if sheet.ensino[i] == True:
            ws.add_image(checked,celula)
        else:
            ws.add_image(unchecked,celula)


#dadosCabecalho()
#ritAulas()
#subTotais()
ritAtividadesEnsino()



wb.save('teste planilha RIT.xlsx')

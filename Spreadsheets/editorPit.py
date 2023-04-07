from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = load_workbook('Anexo 2 - Planilha RIT.xlsx')
ws = wb.active
ws.title = 'Rit'



#FUNÇÃO CRIAR RIT

def dadosCabecalho():
    ws['C3'].value = ''
    ws['C4'].value = ''
    ws['C5'].value = ''
    ws['C6'].value = ''
    ws['C7'].value = ''
    ws['C8'].value = ''
    ws['C9'].value = ''

dadosCabecalho()

wb.save('Anexo 2 - Planilha RIT.xlsx')




  
